from selenium import webdriver
from selenium.webdriver.chrome.options import options
from selenium.webdriver.common.by import by
from selenium.webdriver.support.expected_conditions import _find_elements
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import *
from time import sleep
import openpyxl
import smtplib
import os
from email.message import EmailMessage
import re

class Scrappy:
    def Iniciar(self):
        self.email_usuário()
        self.raspagem_de_dados()
        self.criar_planilha()
        self.enviar_email_cliente()

    def email_usuário(self):

        self.email = input('Digite o email para receber o relatório de valores  dos celulares!\n')
        self.email.lower()
        self.senha = input('Digite a sua senha: ')
        padrao = re.search(r'[a-zA-Z0-9_-]+@[a-zA-Z0-9]+\n.[a-zA-Z]{1,3}$', self.email)
        if padrao:
            print('email válido')

        else:
            print('Digite um email valido!!!')
            self.email_usuário()
    
    def raspagem_de_dados(self):
        chrome_options = Options()
        chrome_options.add_experimental_options('excludeSwitches', ['enable-logging'])
        chrome_options.add_argument('--lang=pr-BR')
        chrome_options.add_argument('--disable-notifications')
        self.driver = webdriver.Chrome(executable_path=os.getcwd()+ os.sep + 'chromedriver.exe', options=chrome_options)
        self.driver.set_windows_size(800, 700)
        self.link = 'https: //telefonesimportados.netfy.app/'
        print(self.driver.title) 
        self.lista_nome_celulares = []
        self.lista_preco_celulares = []
        self.driver.get(self.link)
        sleep(2)

        for p in range(5):
            item = 1
            for i in  range(12):
                lista_nomes = self.driver._find_elements_by_xpath(f'/html/body/div[5]/div[2]/div[1]/div[{item}]/div/h2/a')
                self.lista_nome_celulares.append(lista_nomes[0].text)
                sleep(1)
                lista_precos = self.driver._find_elements_by_xpath(f'//div[{item}/div[@class="single-shop-product" and 1]/div[@class="product-carousel-price" and 2]/ins[1]')
                self.lista_preco_celulares.append(lista_precos[0].text)
                item += 1
                sleep(1)
    try:
        botao_proximo = self.driver._find_element_by_xpath(
            '/html/body/div[5]/div[2]/div[2]/div/div/nav/ul/li/[7]/a'
        )
        botao_proximo.click()
        print(f'\u001b[32{"Navegando para a proxima pagina"}\u001b[0m')
        sleep(2)

    except NoSuchElementException:

        print(f'\u001b[33m{"Não há mais páginas!"}\u001b[0m')
        print(f'\u001b[32m{"Escaneamento Concluido"}\u001b[0m')
        self.driver.quit()

    def criar_planilha(self):
        index = 2
        planilha = openpyxl.Workbook()
        celulares = planilha['Sheet']
        celulares['A1'] = 'Nome'
        celulares['B1'] = 'Preço'
        for nome, preco in zip(self.lista_nome_celulares, self.lista_preco_celulares):
            celulares.cell(column=1, row=index, value=nome)
            celulares.cell(column=2, row=index, value=preco)
            index += 1
        planilha.save("planilha_de_precos.xlsx")
        print(f'\u001b[32{"Planilha criada com successo"}\u001b[0m')

    def enviar_email_cliente(self, endereco, senha):
        endereco = 'Digite o email aqui'
        senha = 'Digite a senha aqui'
        msg = EmailMessage()
        msg['Subject'] = 'Planilha de preços de telefones importado'
        msg['From'] = endereco
        msg['To'] = self.email
        msg = set_content('Olá, a sua planilha chegou!')
        arquivos = ["planilha_de_preços.xlsx"]
        for arquivo in arquivos:
            with open(arquivo, 'rb') as arq:
                dados = arq.read()
                nome_arquivo = arq.nome
            msg.add_attachment(dados, maintype='aplication',  subtype='octet-stream',filename=nome_arquivo)
        
        server = smtplib.SMTP('imap.gmail.com', 587)
        server.ehlo()
        server.starttls()
        server.login(endereco, senha, initial_response_ok=True)
        server.send_message(msg)
        print(f'\u001b[32m{"Enviando email para o destinatário"}\u001b[0m')

        server.quit()


start = Scrappy()
start.Iniciar()